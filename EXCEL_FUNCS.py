import os
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string, get_column_letter
from openpyxl.styles import Border, Side, Alignment, PatternFill,Font
from openpyxl.styles import Protection
from openpyxl.drawing.image import Image
import xlwings
from DATA_FUNCS import *


def ReturnRow(ws)  :  # 返回"产品型号"的行号 与 项目名称
    name =row =None
    for i in range(1 ,100):
        if name==None and ws.cell(i ,2).value.startswith("项目名称"):
            if len(ws.cell(i ,2).value )==5:
                name =""
            else:
                name =ws.cell(i ,2).value[5:]
                if "\n" in name:
                    name =name.split("\n")[0]
        if row==None and ws.cell(i ,8).value=="产品型号":
            row =i
        if name and row:
            break
    return row ,name

def TellRegion(ws ,row):
    end_row = coordinate_from_string(ws.dimensions.split(":")[1])[1]
    tmp_lst=[] #型号列非空行号列表
    reg_name_lst=[] #区域列非空行号列表
    loop_row=row+1
    while loop_row<=end_row:
        if ws.cell(loop_row,8).value!=None:
            tmp_lst.append(loop_row)
        elif ws.cell(loop_row,2).value!=None:
            reg_name_lst.append(loop_row)
        loop_row+=1
    ws_end_row=tmp_lst[-1]
    region_lst=[[]] # [[区域名1，开始行号，结束行号],[区域名2，开始行号，结束行号]]
    for num in tmp_lst:
        if num-1 in region_lst[-1]:
            region_lst[-1].append(num)
        else:
            region_lst.append([num])
    del region_lst[0]
    for grp in region_lst:
        for i in range(len(grp)-2):
            del grp[1]
    for i in range(len(region_lst)):
        for name_row in reg_name_lst:
            if i==0:
                if name_row < region_lst[i][0]:
                    region_lst[i].insert(0,ws.cell(name_row,2).value)
                    reg_name_lst.remove(name_row)
                    break
            else:
                if name_row < region_lst[i][0] and name_row > region_lst[i-1][-1]:
                    region_lst[i].insert(0,ws.cell(name_row,2).value)
                    reg_name_lst.remove(name_row)
                    break
    for grp in region_lst:
        if not isinstance(grp[0], str):
            grp.insert(0,'')
        if len(grp) == 2:
            grp.append(grp[-1])
    return region_lst,ws_end_row

def get_datas(ws,region_lst):
    data_groups=[]
    KVAR_and_A={}
    for reg in region_lst:
        KVAR_and_A[reg[0]]={}
        data_each_reg=[]
        model_and_quantity=[]
        for row in range(reg[1],reg[2]+1):
            product_model=ws.cell(row,8).value.strip() #型号
            quantity=ws.cell(row,13).value #数量
            capacity_statistics(product_model, quantity, KVAR_and_A[reg[0]])#统计容量
            if ws.cell(row,14).value!=None:
                remark=ws.cell(row,14).value.strip() #备注
            else:
                remark=''
            if product_model+remark not in model_and_quantity:
                model_and_quantity.append(product_model+remark)
                if product_model.startswith("KLD-BK"):
                    if "不" in product_model and "电抗" in product_model:
                        BK_v=product_model.split("-")[-1][0:product_model.split("-")[-1].index("V")] #若除电压外的其他地方有"V",需要变更
                    else:
                        BK_v=product_model.split("-")[-1].rstrip("V")
                    BK_v_int=int(BK_v)
                    BK_kv_int=BK_v_int/1000
                    BK_kv=str(BK_kv_int)
                    sys_u_row=row
                    while type(ws.cell(sys_u_row,4)).__name__=='MergedCell':
                        sys_u_row-=1
                    if ws.cell(sys_u_row,4).value.startswith(BK_v) or ws.cell(sys_u_row,4).value.startswith(BK_kv):
                        data_each_reg.append([product_model,remark,quantity,True,BK_v_int])
                    else:
                        data_each_reg.append([product_model, remark, quantity,False,BK_v_int])
                else:
                    data_each_reg.append([product_model, remark, quantity])
            else:
                index=model_and_quantity.index(product_model+remark)
                data_each_reg[index][2]+=quantity
        data_each_reg.sort()
        data_groups.append(data_each_reg)
    return data_groups,KVAR_and_A

def write_datas(data_groups,region_lst,project_name,info_lst): #info_lst:[日期，姓名，代号]
    wb=load_workbook("电能质量--报价模板.xlsx")
    ws=wb["报价单"]
    end_row = coordinate_from_string(ws.dimensions.split(":")[1])[1]
    ws.cell(2,1).value="报价日期："+info_lst[0]
    ws.cell(4, 4).value="NAME："+info_lst[1]
    ws.cell(8,1).value="项目名称:"+project_name
    reg_name_num=0
    loop_row=9 if region_lst[reg_name_num][0]=='' else 10
    products_and_rows={} #存放产品与行数的字典
    for region in data_groups:
        if loop_row!=9:
            ws.cell(loop_row,1).value=region_lst[reg_name_num][0]
        reg_name_num+=1
        loop_row+=1
        seq_num=1
        for data_lst in region:
            ws.cell(loop_row, 1).value = seq_num #填写每个区域的序号
            seq_num+=1
            ws.cell(loop_row,3).value=data_lst[0] #填型号
            if data_lst[1]!="":
                ws.cell(loop_row,4).value=data_lst[1] #若备注了分路，则填写
            ws.cell(loop_row,5).value=data_lst[2] #填数量
            #开始判断型号
            fill_products_and_row(data_lst, products_and_rows, loop_row )
            loop_row+=1
    return wb,products_and_rows,loop_row,end_row

def bulid_form(wb,products_and_rows_splited,coord_str,BK_correspondence,kvar_sign): #有BKJ且没勾选按kvar报则就地分离
    ws = wb["报价单"]
    coord=[coordinate_from_string(coord_str)[1],column_index_from_string(coordinate_from_string(coord_str)[0])]
    ws.cell(coord[0], coord[1]).value = "调价表"
    ws.cell(coord[0]+1,coord[1]).value = "涉及产品"
    ws.cell(coord[0]+1,coord[1]+1).value = "规格"
    ws.cell(coord[0]+1,coord[1]+2).value = "底价"
    ws.cell(coord[0]+1,coord[1]+3).value = "市场价"
    ws.cell(coord[0]+1,coord[1]+4).value = "底价系数"
    ws.cell(coord[0]+1,coord[1]+5).value = "最终底价"
    ws.cell(coord[0]+1,coord[1]+6).value = "调价"

    product_row = coord[0]+2
    product_row_FL = {}
    price_row = {}
    for product_name in products_and_rows_splited:
        if "KLD-BK" in product_name:
            each_row_FL = {}
            sign_less = sign_greater = None
            if "-1-" in product_name:
                for row in products_and_rows_splited[product_name]:
                    each_row_FL[row] = fenlu(ws.cell(row, 4).value, 1)
                    if each_row_FL[row][0] != []:  # 判断是否存在<= 5的分路
                        sign_less = True
                    if each_row_FL[row][1] != []:  # 是否存在>10的分路
                        sign_greater = True

                if "KLD-BKJ" in product_name and not kvar_sign:
                    for row in each_row_FL:
                        for tup in each_row_FL[row]:
                            for saq in tup:
                                if saq!=[]:
                                    if product_name in BK_correspondence:
                                        BK_correspondence[product_name][saq[0]]=Contactor_selection(saq[0],1,int(float(product_name.split("-")[-1].rstrip("V"))*1.732050808))
                                    else:
                                        BK_correspondence[product_name]={saq[0]:Contactor_selection(saq[0],1,int(float(product_name.split("-")[-1].rstrip("V"))*1.732050808))}
                    product_name=product_name.replace("BKJ", "BKS")
                if product_name in product_row_FL:
                    product_row_FL[product_name].update(each_row_FL)
                else:
                    product_row_FL[product_name]=each_row_FL

                ws.cell(product_row, coord[1]).value = product_name  # 调价表内产品名称
                if sign_less == True and sign_greater == True:
                    price_row[product_name] = [product_row, product_row + 1]
                    ws.cell(product_row, coord[1]+1).value = "≤5"
                    ws.cell(product_row, coord[1] + 2).value, ws.cell(product_row,coord[1] + 3).value = get_price_value(product_name, 0)
                    product_row += 1
                    ws.cell(product_row, coord[1]+1).value = "≥10"
                    ws.cell(product_row, coord[1] + 2).value, ws.cell(product_row,coord[1] + 3).value = get_price_value(product_name, 1)
                    ws.merge_cells(f"{get_column_letter(coord[1])}{product_row - 1}:{get_column_letter(coord[1])}{product_row}")
                elif sign_less == True:
                    price_row[product_name] = [product_row, None]
                    ws.cell(product_row, coord[1]+1).value = "≤5"
                    ws.cell(product_row, coord[1] + 2).value, ws.cell(product_row,coord[1] + 3).value = get_price_value(product_name, 0)
                elif sign_greater == True:
                    price_row[product_name] = [None, product_row]
                    ws.cell(product_row, coord[1]+1).value = "≥10"
                    ws.cell(product_row, coord[1] + 2).value, ws.cell(product_row,coord[1] + 3).value = get_price_value(product_name, 1)
            else:
                product_name_lst=product_name.split("-")
                sys_v_int=int(product_name.split("-")[-1].rstrip("V"))
                if product_name_lst[1]=="BKS14" and sys_v_int > 400 and sys_v_int <= 480 or sys_v_int == 400: #主要针对40kvar
                    sign_0=sign_1=sign_2=sign_3=False
                    for row in products_and_rows_splited[product_name]:
                        each_row_FL[row] = fenlu(ws.cell(row, 4).value, 3,True)
                        if each_row_FL[row][0] != []:  # 是否存在<25的分路
                            sign_0 = True
                        if each_row_FL[row][1] != []:
                            sign_1 = True
                        if each_row_FL[row][2] != []:
                            sign_2 = True
                        if each_row_FL[row][3] != []:
                            sign_3 = True

                    if "KLD-BKJ" in product_name and not kvar_sign:
                        for row in each_row_FL:
                            for tup in each_row_FL[row]:
                                for saq in tup:
                                    if saq != []:
                                        if product_name in BK_correspondence:
                                            BK_correspondence[product_name][saq[0]] = Contactor_selection(saq[0], 3, int(product_name.split("-")[-1].rstrip("V")))
                                        else:
                                            BK_correspondence[product_name] = {saq[0]: Contactor_selection(saq[0], 3, int(product_name.split("-")[-1].rstrip("V")))}
                        product_name = product_name.replace("BKJ", "BKS")
                    if product_name in product_row_FL:
                        product_row_FL[product_name].update(each_row_FL)
                    else:
                        product_row_FL[product_name]=each_row_FL

                    ws.cell(product_row, coord[1]).value = product_name  # 调价表内产品名称
                    start_row=product_row
                    price_row[product_name]=[]
                    if sign_0 == True:
                        ws.cell(product_row, coord[1]+1).value = "容量<25"
                        ws.cell(product_row, coord[1] + 2).value, ws.cell(product_row,coord[1] + 3).value = get_price_value(product_name, 0)
                        price_row[product_name].append(product_row)
                        product_row+=1
                    else:
                        price_row[product_name].append(None)
                    if sign_1 == True:
                        ws.cell(product_row, coord[1]+1).value = "25≤容量≤30"
                        ws.cell(product_row, coord[1] + 2).value, ws.cell(product_row,coord[1] + 3).value = get_price_value(product_name, 1)
                        price_row[product_name].append(product_row)
                        product_row+=1
                    else:
                        price_row[product_name].append(None)
                    if sign_2 == True:
                        ws.cell(product_row, coord[1]+1).value = "30<容量<50"
                        ws.cell(product_row, coord[1] + 2).value, ws.cell(product_row,coord[1] + 3).value = get_price_value(product_name, 2)
                        price_row[product_name].append(product_row)
                        product_row+=1
                    else:
                        price_row[product_name].append(None)
                    if sign_3 == True:
                        ws.cell(product_row, coord[1]+1).value = "容量≥50"
                        ws.cell(product_row, coord[1] + 2).value, ws.cell(product_row,coord[1] + 3).value = get_price_value(product_name, 3)
                        price_row[product_name].append(product_row)
                        product_row+=1
                    else:
                        price_row[product_name].append(None)
                    product_row-=1
                    ws.merge_cells(f"{get_column_letter(coord[1])}{start_row}:{get_column_letter(coord[1])}{product_row}")
                else:
                    for row in products_and_rows_splited[product_name]:
                        each_row_FL[row] = fenlu(ws.cell(row, 4).value, 3) #此时fenlu返回的不是元组
                        if "KLD-BKJ" in product_name and not kvar_sign:
                            for row in each_row_FL:
                                for tup in each_row_FL[row]:
                                    for saq in tup:
                                        if saq != []:
                                            if product_name in BK_correspondence:
                                                BK_correspondence[product_name][saq[0]] = Contactor_selection(saq[0], 3, int(product_name.split("-")[-1].rstrip("V")))
                                            else:
                                                BK_correspondence[product_name] = {saq[0]: Contactor_selection(saq[0], 3, int(product_name.split("-")[-1].rstrip("V")))}
                            product_name = product_name.replace("BKJ", "BKS")
                    if product_name in product_row_FL:
                        product_row_FL[product_name].update(each_row_FL)
                    else:
                        product_row_FL[product_name]=each_row_FL
                    ws.cell(product_row, coord[1]).value = product_name  # 调价表内产品名称
                    price_row[product_name] = [product_row]
                    ws.cell(product_row, coord[1] + 2).value, ws.cell(product_row,coord[1] + 3).value = get_price_value(product_name, 0)
                    ws.merge_cells(f"{get_column_letter(coord[1])}{product_row}:{get_column_letter(coord[1]+1)}{product_row}")
        elif "AMS" in product_name:
            ws.cell(product_row, coord[1]).value = product_name
            product_range = []
            product_range_info = []
            product_row_FL[product_name] = {}
            for row in products_and_rows_splited[product_name]:
                svg_lst = ws.cell(row, 3).value.split("-")
                capacity_apf = int(svg_lst[1].lstrip("AMS"))
                price_info = get_price_value(product_name, capacity_apf)#得到（底价，面价，规格范围）
                if price_info[2] in product_range:
                    product_row_FL[product_name][row] = [product_range.index(price_info[2]),price_info[-1]]
                else:
                    product_range.append(price_info[2])
                    product_range_info.append(price_info)
                    product_row_FL[product_name][row] = [product_range.index(price_info[2]),price_info[-1]]
            price_row[product_name] = []
            for i in range(len(product_range)):
                ws.cell(product_row, coord[1] + 2).value, ws.cell(product_row, coord[1] + 3).value, ws.cell(
                    product_row, coord[1] + 1).value = product_range_info[i][0],product_range_info[i][1],product_range_info[i][2]
                price_row[product_name].append(product_row)
                product_row += 1
            ws.merge_cells(f"{get_column_letter(coord[1])}{product_row - len(product_range)}:{get_column_letter(coord[1])}{product_row - 1}")
            product_row -= 1
        elif "KLD-SVG" in product_name:
            ws.cell(product_row, coord[1]).value = product_name
            svg_range=[]
            svg_range_info=[]
            product_row_FL[product_name]={}
            for row in products_and_rows_splited[product_name]:
                svg_lst=ws.cell(row, 3).value.split("-")
                capacity_svg = int(svg_lst[2])
                price_info=get_price_value(product_name,capacity_svg)
                if price_info[2] in svg_range:
                    product_row_FL[product_name][row]=[svg_range.index(price_info[2]),price_info[-1]]
                else:
                    svg_range.append(price_info[2])
                    svg_range_info.append(price_info)
                    product_row_FL[product_name][row]=[svg_range.index(price_info[2]),price_info[-1]]
            price_row[product_name] = []
            for i in range(len(svg_range)):
                ws.cell(product_row, coord[1] + 2).value, ws.cell(product_row, coord[1] + 3).value, ws.cell(product_row, coord[1] + 1).value =svg_range_info[i][0],svg_range_info[i][1],svg_range_info[i][2]
                price_row[product_name].append(product_row)
                product_row += 1
            ws.merge_cells(f"{get_column_letter(coord[1])}{product_row - len(svg_range)}:{get_column_letter(coord[1])}{product_row-1}")
            product_row -= 1
        elif "MK-H1" in product_name:
            each_row_QK = {}
            sign_0 = sign_1 = sign_2 = None
            for row in products_and_rows_splited[product_name]:
                capacity_MK_H1 = int(ws.cell(row, 3).value.split("-")[3])
                if capacity_MK_H1 < 60 and capacity_MK_H1>=30:
                    each_row_QK[row] = 0
                    sign_0 = True
                elif capacity_MK_H1 >=60 and capacity_MK_H1 <=134:
                    each_row_QK[row] = 1
                    sign_1 = True
                elif capacity_MK_H1 > 134:
                    each_row_QK[row] = 2
                    sign_2 = True
            product_row_FL[product_name] = each_row_QK
            ws.cell(product_row, coord[1]).value = product_name
            sign_multi=0
            price_row[product_name] = [None, None, None]
            if sign_0==True:
                price_row[product_name][0]=product_row
                ws.cell(product_row, coord[1]+1).value = "30≤容量<60"
                ws.cell(product_row, coord[1] + 2).value,ws.cell(product_row, coord[1] + 3).value=get_price_value(product_name,0)
                sign_multi+=1
            if sign_1==True:
                if sign_multi!=0:
                    product_row += 1
                price_row[product_name][1] = product_row
                ws.cell(product_row, coord[1] + 1).value = "60≤容量≤134"
                ws.cell(product_row, coord[1] + 2).value,ws.cell(product_row, coord[1] + 3).value=get_price_value(product_name,1)
                sign_multi+=1
            if sign_2==True:
                if sign_multi!=0:
                    product_row += 1
                price_row[product_name][2] = product_row
                ws.cell(product_row, coord[1] + 1).value = ">134"
                ws.cell(product_row, coord[1] + 2).value,ws.cell(product_row, coord[1] + 3).value=get_price_value(product_name,2)
                sign_multi+=1
            if sign_multi!=1:
                ws.merge_cells(f"{get_column_letter(coord[1])}{product_row - sign_multi +1}:{get_column_letter(coord[1])}{product_row}")
        elif "FD" in product_name and "-H" in product_name:
            each_row_QK = {}
            sign_0 = sign_1 = sign_2 = None
            for row in products_and_rows_splited[product_name]:
                capacity_FD_H = int(ws.cell(row, 3).value.split("-")[3])
                if capacity_FD_H < 200 and capacity_FD_H>=80:
                    each_row_QK[row] = 0
                    sign_0 = True
                elif capacity_FD_H >=200 and capacity_FD_H <=402:
                    each_row_QK[row] = 1
                    sign_1 = True
                elif capacity_FD_H > 402:
                    each_row_QK[row] = 2
                    sign_2 = True
            product_row_FL[product_name] = each_row_QK
            ws.cell(product_row, coord[1]).value = product_name
            sign_multi=0
            price_row[product_name] = [None, None, None]
            if sign_0==True:
                price_row[product_name][0] = product_row
                ws.cell(product_row, coord[1]+1).value = "80≤容量<200"
                ws.cell(product_row, coord[1] + 2).value,ws.cell(product_row, coord[1] + 3).value=get_price_value(product_name,0)
                sign_multi+=1
            if sign_1==True:
                if sign_multi!=0:
                    product_row += 1
                price_row[product_name][1] = product_row
                ws.cell(product_row, coord[1] + 1).value = "200≤容量≤402"
                ws.cell(product_row, coord[1] + 2).value,ws.cell(product_row, coord[1] + 3).value=get_price_value(product_name,1)
                sign_multi+=1
            if sign_2==True:
                if sign_multi!=0:
                    product_row += 1
                price_row[product_name][2] = product_row
                ws.cell(product_row, coord[1] + 1).value = ">402"
                ws.cell(product_row, coord[1] + 2).value,ws.cell(product_row, coord[1] + 3).value=get_price_value(product_name,2)
                sign_multi+=1
            if sign_multi!=1:
                ws.merge_cells(f"{get_column_letter(coord[1])}{product_row - sign_multi +1}:{get_column_letter(coord[1])}{product_row}",0)
        else:
            product_row_FL[product_name] = products_and_rows_splited[product_name]
            ws.cell(product_row, coord[1]).value = product_name  # 调价表内产品名称
            ws.cell(product_row, coord[1] + 2).value,ws.cell(product_row, coord[1] + 3).value=get_price_value(product_name)
            ws.merge_cells(f"{get_column_letter(coord[1])}{product_row}:{get_column_letter(coord[1]+1)}{product_row}")
            ws.cell(product_row, coord[1] + 2).value, ws.cell(product_row,coord[1] + 3).value = get_price_value(product_name)
            price_row[product_name] = [product_row]
        product_row += 1
    #开始单独处理接触器（填入名称与价格）
    tmp_lst=[]
    for p in BK_correspondence:
        if isinstance(BK_correspondence[p],dict):
            for kvar in BK_correspondence[p]:
                jname=BK_correspondence[p][kvar]
                if jname not in tmp_lst:
                    tmp_lst.append(jname)
    for j in tmp_lst:
        ws.cell(product_row, coord[1]).value = j
        ws.merge_cells(f"{get_column_letter(coord[1])}{product_row}:{get_column_letter(coord[1] + 1)}{product_row}")
        ws.cell(product_row, coord[1] + 2).value, ws.cell(product_row,coord[1] + 3).value = get_price_value(j)
        price_row[j] = [product_row]
        product_row += 1

    ws_adj_end_row = product_row-1
    return ws,product_row_FL,price_row,ws_adj_end_row,coord,BK_correspondence

def set_formula(ws,products_and_rows,product_row_FL,price_row,BK_correspondence,ws_adj_end_row,coord,sign):
    adj_column_5=get_column_letter(coord[1]+5) #最终底价
    adj_column_6=get_column_letter(coord[1]+6) #调价
    # 设定调价表内公式
    for row in range(coord[0]+2,ws_adj_end_row+1):
        ws.cell(row,coord[1]+4).value=1
        ws.cell(row,coord[1]+5).value=f"={get_column_letter(coord[1]+2)}{row}*{get_column_letter(coord[1]+4)}{row}"
    #设定报价单内公式
    for project in products_and_rows:
        dynamic_sign = static_sign=False
        if not sign:#若未勾选按kvar报价
            if "BKT" in project or "BKC" in project:
                project_nobkt = project
                project = project.replace("BKT", "BKS")
                project = project.replace("BKC", "BKS")
                dynamic_sign = True
            elif "BKJ" in project:
                project_nobkt = project
                project = project.replace("BKJ", "BKS")
                static_sign=True
            else:
                project_nobkt = project
        else:
            if "BKT" in project or "BKC" in project or "BKJ" in project:
                for row in products_and_rows[project]:
                    tup = product_row_FL[project][row] #([[20, 2]], [], [[40, 1]], [[50, 4]])
                    eqt_tup=''
                    spe_num=0
                    for saq_lst in tup:
                        if saq_lst == []:
                            spe_num+=1
                        else:
                            for saq in saq_lst:
                                eqt_tup+=f'+{adj_column_6}{price_row[project][spe_num]}*{saq[0]}*{saq[1]}'
                            spe_num+=1
                    ws.cell(row, 7).value = f'={eqt_tup.lstrip("+")}'
                continue
            else:
                project_nobkt = project
        for row in products_and_rows[project_nobkt]:
            if type(product_row_FL[project]) is dict:  # BK、SVG、APF、MK-H3、FD-H
                # if "不" in project_nobkt and "电抗" in project_nobkt: #新版本移除此功能
                #     project_nobkt=project_nobkt[0:project_nobkt.index("V")+1]
                capacity=0 #定义一下防止标黄
                if type(product_row_FL[project][row]) is tuple:
                    tup = product_row_FL[project][row]  # 分路元组
                    eqt_tup_lst=[]
                    eqt_tup=eqt_tsc=""
                    i=0
                    for saq_list in tup:
                        if saq_list==[]:
                            i+=1
                            continue
                        else:
                            for saq in saq_list:
                                eqt_tup_lst.append(f"{adj_column_5}{price_row[project][i]}*{saq[0]}*{saq[1]}")
                                if dynamic_sign == True:
                                    FL_quantity=saq[1]
                                    if "-1-" in project_nobkt:
                                        FL_quantity/=3
                                    eqt_tsc += f"+{adj_column_5}{price_row[BK_correspondence[project_nobkt]][0]}*{FL_quantity}"
                                elif static_sign == True:
                                    FL_quantity=saq[1]
                                    eqt_tsc += f"+{adj_column_5}{price_row[BK_correspondence[project_nobkt][saq[0]]][0]}*{FL_quantity}"
                        i+=1
                    for e in eqt_tup_lst:
                        e_adj=e.replace(f"{adj_column_5}", f"{adj_column_6}")
                        eqt_tup+=f"+IF({e_adj},{e_adj},{e})"
                    eqt_tup=eqt_tup.lstrip("+")
                    if dynamic_sign == True or static_sign == True:
                        eqt_tsc=eqt_tsc.lstrip("+")
                        eqt_tsc_adj=eqt_tsc.replace(f"{adj_column_5}", f"{adj_column_6}")
                        ws.cell(row, 7).value = f"={eqt_tup}+IF({eqt_tsc_adj},{eqt_tsc_adj},{eqt_tsc})"
                    else:
                        ws.cell(row, 7).value = f"={eqt_tup}"
                elif "KLD-SVG" in project:
                    if product_row_FL[project][row][1]==1:
                        capacity = ws.cell(row, 3).value.split("-")[2]
                    elif product_row_FL[project][row][1]==0:
                        capacity="1"
                    eqt = f"{capacity}*{adj_column_5}{price_row[project][product_row_FL[project][row][0]]}"
                    eqt_adj = eqt.replace(f"{adj_column_5}",f"{adj_column_6}")
                    ws.cell(row, 7).value = f"=IF({eqt_adj},{eqt_adj},{eqt})"
                elif "AMS" in project:
                    if product_row_FL[project][row][1]==1:
                        capacity = ws.cell(row, 3).value.split("-")[1].lstrip("AMS")
                    elif product_row_FL[project][row][1]==0:
                        capacity="1"
                    eqt = f"{capacity}*{adj_column_5}{price_row[project][product_row_FL[project][row][0]]}"
                    eqt_adj = eqt.replace(f"{adj_column_5}",f"{adj_column_6}")
                    ws.cell(row, 7).value = f"=IF({eqt_adj},{eqt_adj},{eqt})"
                else:  # SVG、APF、MK-H3、FD-H
                    capacity = 0
                    if "MK-H1" in project:
                        capacity = ws.cell(row, 3).value.split("-")[3]
                    elif "FD" in project and "-H" in project:
                        capacity = ws.cell(row, 3).value.split("-")[3]
                    eqt = f"{capacity}*{adj_column_5}{price_row[project][product_row_FL[project][row]]}"
                    eqt_adj = eqt.replace(f"{adj_column_5}",f"{adj_column_6}")
                    ws.cell(row, 7).value = f"=IF({eqt_adj},{eqt_adj},{eqt})"
            else:  # 其他产品
                eqt_others = f"{adj_column_5}{price_row[project][0]}"
                eqt_others_adj = eqt_others.replace(f"{adj_column_5}",f"{adj_column_6}")
                ws.cell(row, 7).value = f"=IF({eqt_others_adj},{eqt_others_adj},{eqt_others})"

def set_format(ws,coord,ws_adj_end_row):# 设置调价表格式
    ws.merge_cells(f"{get_column_letter(coord[1])}{coord[0]}:{get_column_letter(coord[1]+6)}{coord[0]}")
    ws.cell(coord[0],coord[1]).alignment = Alignment(horizontal='center', vertical='center')
    borders = ws[f"{get_column_letter(coord[1])}{coord[0]}:{get_column_letter(coord[1]+6)}{ws_adj_end_row}"]
    thin_boader = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),bottom=Side(style='thin'))
    for j in borders:
        for i in j:
            i.border = thin_boader
    ws.column_dimensions[get_column_letter(coord[1])].width = 25
    ws.column_dimensions[get_column_letter(coord[1]+1)].width = 10
    for c in range(2,7):
        ws.column_dimensions[get_column_letter(coord[1] + c)].width = 10
    for r in range(coord[0]+2, ws_adj_end_row + 1):
        ws.cell(r, coord[1]).alignment = Alignment(horizontal='left', vertical='center')

def statistical_table(ws,ws_adj_end_row,coord,KVAR_and_A):
    coord[1]+=8
    #设定统计表内总容量单元格的颜色
    color_fill = PatternFill(fill_type='solid', fgColor="B2B2B2")
    #建立容量统计表格
    first_stat_row=stat_row=3
    stat_column=coord[1]+2
    ws.cell(stat_row,coord[1]).value="容量统计"
    names=[]
    stat_row+=1
    ws.cell(stat_row,coord[1]).value="容量种类\区域"
    ws.merge_cells(f"{get_column_letter(coord[1])}{stat_row}:{get_column_letter(coord[1]+1)}{stat_row}")#合并"容量种类\区域"
    for reg in KVAR_and_A:#在表内填写区域名称，同时合并整理各区域字典的key
        ws.cell(stat_row,stat_column).value=reg
        stat_column+=1
        for name in KVAR_and_A[reg]:
            if name not in names:
                names.append(name)
    ws.cell(stat_row,stat_column).value="产品总容量"
    end_column=stat_column#容量统计表最右列
    end_row=stat_row
    amount=len(names)
    amount_kvar=0
    if "APF" in names:
        stat_row+=1
        ws.cell(stat_row,coord[1]).value="滤波(A)"
        ws.merge_cells(f"{get_column_letter(coord[1])}{stat_row}:{get_column_letter(coord[1]+1)}{stat_row}")#合并"滤波(A)"
        names.pop(names.index("APF"))
        APF_column=coord[1]+2
        for reg in KVAR_and_A:
            if "APF" in KVAR_and_A[reg]:
                ws.cell(stat_row,APF_column).value=KVAR_and_A[reg]["APF"]
            else:
                ws.cell(stat_row,APF_column).value="/"
            APF_column+=1
        end_row=stat_row#若无补偿，则为最下行
        ws.cell(stat_row,end_column).fill=color_fill#APF总容量单元格颜色
    if names!=[]:
        amount_kvar=len(names)
        stat_row+=1
        ws.cell(stat_row, coord[1]).value = "补偿(kvar)"
        for name in names:
            column=coord[1]+1
            ws.cell(stat_row,column).value=name
            column+=1
            for reg in KVAR_and_A:
                if name in KVAR_and_A[reg]:
                    ws.cell(stat_row, column).value = KVAR_and_A[reg][name]
                else:
                    ws.cell(stat_row, column).value = "/"
                column+=1
            stat_row+=1
        ws.cell(stat_row,coord[1] ).value = "区域总补偿容量"
        end_row=stat_row#最低行
        ws.cell(end_row,end_column).fill=color_fill
        #设置剩下的单元各合并
        ws.merge_cells(f"{get_column_letter(coord[1])}{stat_row-len(names)}:{get_column_letter(coord[1])}{stat_row-1}")#合并“补偿(kvar)”
        ws.merge_cells(f"{get_column_letter(coord[1])}{stat_row}:{get_column_letter(coord[1] + 1)}{stat_row}")#合并“区域总补偿容量”
    ws.merge_cells(f"{get_column_letter(coord[1])}{first_stat_row}:{get_column_letter(end_column)}{first_stat_row}")#合并“容量统计”

    #设定统计表公式
    for i in range(amount+1):
        ws.cell(first_stat_row+2+i,end_column).value=f"=SUM({get_column_letter(coord[1]+2)}{first_stat_row+2+i}:{get_column_letter(end_column-1)}{first_stat_row+2+i})"#“产品总容量”的公式
    for i in range(len(KVAR_and_A)+1):
        ws.cell(end_row, coord[1]+2+i).value =f"=SUM({get_column_letter(coord[1]+2+i)}{end_row-amount_kvar}:{get_column_letter(coord[1]+2+i)}{end_row-1})"#"区域总补偿容量"公式

    #设定统计表边框
    borders = ws[f"{get_column_letter(coord[1])}{first_stat_row}:{get_column_letter(end_column)}{end_row}"]
    thin_boader = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),bottom=Side(style='thin'))
    for j in borders:
        for i in j:
            i.border = thin_boader
            i.alignment = Alignment(horizontal='center', vertical='center')

def Lock_and_protect(ws,coord,ws_adj_end_row):
    #设置单元格锁定与工作表保护
    protection = Protection(locked=False,hidden=False)
    for colunm in [coord[1]+4,coord[1]+6]:
        for row in range(coord[0]+2,ws_adj_end_row+1):
            ws.cell(row, colunm).protection = protection
    # ws.protection.password = 'password123'

def save_file(path_open,code,project_name,wb,str,suffix_sign,suffix_content):
    path_save = path_open.rstrip(path_open.split("/").pop(-1))
    file_save_path=f"{path_save}{code}{project_name}--W2W2({str}).xlsx"
    if suffix_sign:
        while os.path.exists(file_save_path):
            file_save_path=file_save_path.rstrip(".xlsx")+f"{suffix_content}.xlsx"
    wb.save(file_save_path)
    return file_save_path

def delete_rows(path,startrow,endrow):#中途不打开excel
    app = xlwings.App(visible=False)
    wb = app.books.open(path)
    sheet = wb.sheets[0]
    sheet.range(f'{startrow}:{endrow}').api.EntireRow.Delete()
    wb.save(path)
    app.quit()