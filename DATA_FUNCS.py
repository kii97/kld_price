from openpyxl import load_workbook
from openpyxl.utils.cell import coordinate_from_string

def deal_with_dic(dic,key,value):#用于函数capacity_statistics
    if key in dic:
        dic[key]+=value
    else:
        dic[key]=value#yo

def capacity_statistics(product_model,quantity,dic):#容量统计
    if "KLD-BK" in product_model:
        product_model_lst = product_model.split("-")
        deal_with_dic(dic, product_model_lst[1], int(product_model_lst[2]) * quantity)
    elif "SBKS" in product_model:
        product_model_lst = product_model.split("-")
        deal_with_dic(dic, product_model_lst[1], int(product_model_lst[2]) * quantity)
    elif "KLD-SVG" in product_model:
        product_model_lst = product_model.split("-")
        deal_with_dic(dic, product_model_lst[1], int(product_model_lst[2]) * quantity)
    elif "AMS" in product_model:
        product_model_lst = product_model.split("-")
        deal_with_dic(dic, "APF", int(product_model_lst[1].lstrip("AMS")) * quantity)
    elif "MK-H" in product_model:  # KLD-MK-H3-123-12/|3KV，高压部分只捕捉电容器，如果配置单只有电抗器则会失效
        product_model_lst = product_model.split("-")
        deal_with_dic(dic, f"{product_model_lst[1]}-{product_model_lst[2]}",int(product_model_lst[3]) * quantity)
    elif "MK" in product_model:  # KLD-MK-25-480V,同高压。如果配置单只有电抗器则会失效
        product_model_lst = product_model.split("-")
        deal_with_dic(dic, "MK", float(product_model_lst[2]) * quantity)

def Contactor_selection(kvar:float,phase:int,vol:int): #接触器选型我就偷懒不写在数据库里了
    index=[{3:{
        10:0,12.5:1,15:2,20:3,25:4,30:5,40:6,50:7,60:8,80:9
    },1:{
        5:2,10:5,15:7,20:8
    }},{3:{
        5:0,7.5:1,10:2,12.5:3,15:4,20:5,25:6,30:7,40:8,50:9
    },1:{
        5:4,10:7,15:9
    }}]
    NYKC2=["NYKC2-1820C/3P AC230V","NYKC2-2520C/3P AC230V","NYKC2-3220C/3P AC230V","NYKC2-4021C/3P AC230V","NYKC2-5021C/3P AC230V","NYKC2-6521C/3P AC230V","NYKC2-8021C/3P AC230V","NYKC2-9521C/3P AC230V","NYKC2-11510C/3P AC230V","NYKC2-15010C/3P AC230V"]
    try:
        if vol>=380 and vol<=690:
            return NYKC2[index[0][phase][kvar]]
        elif vol>=208 and vol<=230:
            return NYKC2[index[1][phase][kvar]]
    except:
        return None

def checkBox_BKSTSC_ischecked(data_groups):
    for region in data_groups:
        lst_tsc={3:{},1:{}}
        lst_cs={3:{},1:{}}
        lst_c2={3:{},1:{}}
        for product in region:
            #处理可控硅投切的情况
            if "KLD-BKT" in product[0]:
                product[0]=product[0].replace("BKT","BKS")
                if product[3]==True:
                    lst_FL = fenlu(product[1],3)
                    for spec in lst_FL:
                        if spec != []:
                            for saq in spec:
                                if product[4] in lst_tsc[3]:
                                    if saq[0] in lst_tsc[3][product[4]]:
                                        lst_tsc[3][product[4]][saq[0]]+=saq[1]*product[2]
                                    else:
                                        lst_tsc[3][product[4]][saq[0]]=saq[1]*product[2]
                                else:
                                    lst_tsc[3][product[4]]={saq[0]:saq[1]*product[2]}
                else:
                    lst_FL = fenlu(product[1],1)
                    for spec in lst_FL:
                        if spec != []:#[[15, 3], [10, 6]]
                            for saq in spec:#[15, 3]
                                if product[4] in lst_tsc[1]:
                                    if saq[0]*3 in lst_tsc[1][product[4]]:
                                        lst_tsc[1][product[4]][saq[0]*3]+=int(saq[1]*product[2]/3)#防止变成浮点数
                                    else:
                                        lst_tsc[1][product[4]][saq[0]*3]=int(saq[1]*product[2]/3)
                                else:
                                    lst_tsc[1][product[4]]={saq[0]*3:int(saq[1]*product[2]/3)}
            #处理复合开关投切的情况
            if "KLD-BKC" in product[0]:
                product[0]=product[0].replace("BKC","BKS")
                if product[3]==True:
                    lst_FL = fenlu(product[1],3)
                    for spec in lst_FL:
                        if spec != []:
                            for saq in spec:
                                if product[4] in lst_cs[3]:
                                    if saq[0] in lst_cs[3][product[4]]:
                                        lst_cs[3][product[4]][saq[0]]+=saq[1]*product[2]
                                    else:
                                        lst_cs[3][product[4]][saq[0]]=saq[1]*product[2]
                                else:
                                    lst_cs[3][product[4]]={saq[0]:saq[1]*product[2]}
                else:
                    lst_FL = fenlu(product[1],1)
                    for spec in lst_FL:
                        if spec != []:#[[15, 3], [10, 6]]
                            for saq in spec:#[15, 3]
                                if product[4] in lst_cs[1]:
                                    if saq[0]*3 in lst_cs[1][product[4]]:
                                        lst_cs[1][product[4]][saq[0]*3]+=int(saq[1]*product[2]/3)#防止变成浮点数
                                    else:
                                        lst_cs[1][product[4]][saq[0]*3]=int(saq[1]*product[2]/3)
                                else:
                                    lst_cs[1][product[4]]={saq[0]*3:int(saq[1]*product[2]/3)}
            #处理接触器投切的情况
            if "KLD-BKJ" in product[0]:
                product[0]=product[0].replace("BKJ","BKS")
                if product[3]==True:
                    lst_FL = fenlu(product[1],3)
                    for spec in lst_FL:
                        if spec != []:
                            for saq in spec:
                                if product[4] in lst_c2[3]:
                                    if saq[0] in lst_c2[3][product[4]]:
                                        lst_c2[3][product[4]][saq[0]]+=saq[1]*product[2]
                                    else:
                                        lst_c2[3][product[4]][saq[0]]=saq[1]*product[2]
                                else:
                                    lst_c2[3][product[4]]={saq[0]:saq[1]*product[2]}
                else:
                    lst_FL = fenlu(product[1],1)
                    for spec in lst_FL:
                        if spec != []:
                            for saq in spec:
                                if product[4] in lst_c2[1]:
                                    if saq[0] in lst_c2[1][product[4]]:
                                        lst_c2[1][product[4]][saq[0]]+=int(saq[1]*product[2])#防止变成浮点数
                                    else:
                                        lst_c2[1][product[4]][saq[0]]=int(saq[1]*product[2])
                                else:
                                    lst_c2[1][product[4]]={saq[0]:int(saq[1]*product[2])}
        if lst_tsc[3] !={}:
            for voltage in lst_tsc[3]:
                for key in lst_tsc[3][voltage]:
                    if voltage==400:
                        region.append([f"KLD-TSC{key}-3-2010X","",lst_tsc[3][voltage][key]])
                    else:
                        region.append([f"KLD-TSC{key}-3-2010X-{voltage}V","",lst_tsc[3][voltage][key]])
        if lst_tsc[1] != {}:
            for voltage in lst_tsc[1]:
                for key in lst_tsc[1][voltage]:
                    if voltage==230:
                        region.append([f"KLD-TSC{key}-1-2010X","",lst_tsc[1][voltage][key]])
                    else:
                        region.append([f"KLD-TSC{key}-1-2010X-{voltage}V","",lst_tsc[1][voltage][key]])
        if lst_cs[3] !={}:
            for voltage in lst_cs[3]:
                for key in lst_cs[3][voltage]:
                    if voltage==400:
                        region.append([f"KLD-CS{key}-3-2010","",lst_cs[3][voltage][key]])
                    else:
                        region.append([f"KLD-CS{key}-3-2010-{voltage}V","",lst_cs[3][voltage][key]])
        if lst_cs[1] != {}:
            for voltage in lst_cs[1]:
                for key in lst_cs[1][voltage]:
                    if voltage==230:
                        region.append([f"KLD-CS{key}-1-2010","",lst_cs[1][voltage][key]])
                    else:
                        region.append([f"KLD-CS{key}-1-2010-{voltage}V","",lst_cs[1][voltage][key]])
        if lst_c2[3] !={}:
            for voltage in lst_c2[3]:
                for key in lst_c2[3][voltage]:
                    region.append([Contactor_selection(key,3,voltage),"",lst_c2[3][voltage][key]])
        if lst_c2[1] !={}:
            for voltage in lst_c2[1]:
                for key in lst_c2[1][voltage]:
                    region.append([Contactor_selection(key,1,voltage*1.7320508),"",lst_c2[1][voltage][key]]) #这里的系统电压是相电压*1.7320508,所以可能有误差生成不了
        #合并三相单相中规格一致的接触器数量
    for reg in data_groups:
        add_dic={}
        del_lst=[]
        for pdt_info in reg:
            if pdt_info[0].startswith("NYKC2"):
                del_lst.append(pdt_info)
                if pdt_info[0] in add_dic:
                    add_dic[pdt_info[0]] += pdt_info[2]
                else:
                    add_dic[pdt_info[0]]=pdt_info[2]
        for d in del_lst:
            reg.remove(d)
        for p in add_dic:
            reg.append([p,"",add_dic[p]])

def fill_in_dic(model_name_str,products_and_rows,loop_row):#用于fill_products_and_row函数
    if model_name_str not in products_and_rows:
        products_and_rows[model_name_str] = [loop_row]
    else:
        products_and_rows[model_name_str].append(loop_row)

def fill_products_and_row(data_lst,products_and_rows,loop_row):
    if data_lst[0].startswith("KLD"):
        model_lst=data_lst[0].split("-")
        if model_lst[1].startswith("BK"):
            if data_lst[3]==True: #共补
                model_name_str=f"KLD-{model_lst[1]}-{model_lst[3]}"
                fill_in_dic(model_name_str,products_and_rows,loop_row)
            else:
                model_name_str = f"KLD-{model_lst[1]}-1-{model_lst[3]}"
                fill_in_dic(model_name_str,products_and_rows,loop_row)
        elif model_lst[1].startswith("AMS"):#KLD-AMS100-4L-400V-RD
            model_name_str = f"KLD-AMS-{model_lst[2]}-{model_lst[3]}-{model_lst[4]}"
            fill_in_dic(model_name_str,products_and_rows,loop_row)
        elif model_lst[1].startswith("SVG"):#KLD-SVG-100-4L-400V-RD
            model_name_str = f"KLD-SVG-{model_lst[3]}-{model_lst[4]}-{model_lst[5]}"
            fill_in_dic(model_name_str,products_and_rows,loop_row)
        elif model_lst[1].startswith("TSC"):#KLD-TSC20-3-2010A-690V
            model_lst=data_lst[0].split("-")
            if len(model_lst)==4:
                model_name_str=f"KLD-TSC-{model_lst[2]}-2010X"
                fill_in_dic(model_name_str,products_and_rows,loop_row)
            else:
                model_name_str = f"KLD-TSC-{model_lst[2]}-2010X-{model_lst[4]}"
                fill_in_dic(model_name_str, products_and_rows,loop_row)
        elif model_lst[1].startswith("MK") and model_lst[2].startswith("H1"):#KLD-MK-H1-容量-7.2√3KV
            model_name_str=f"KLD-MK-H1-{model_lst[4]}"
            fill_in_dic(model_name_str, products_and_rows,loop_row)
        elif model_lst[1].startswith("FD") and model_lst[2].startswith("H"):#KLD-FD6-H3-容量-6.6KV-L
            if model_lst[2].startswith("H3"):
                model_name_str=f"KLD-{model_lst[1]}-H3-{model_lst[4]}-{model_lst[5]}"
                fill_in_dic(model_name_str, products_and_rows,loop_row)
        elif model_lst[1].startswith("FC"):#KLD-FC6-H1-容量-6.6KV-L
            if model_lst[2].startswith("H1"):
                model_name_str=f"KLD-{model_lst[1]}-H1-{model_lst[4]}-{model_lst[5]}"
                fill_in_dic(model_name_str, products_and_rows,loop_row)
        else:
            fill_in_dic(data_lst[0], products_and_rows,loop_row)
    else:
        fill_in_dic(data_lst[0], products_and_rows,loop_row)

def copy_lst_from_dic(lst):#用于split_bkt函数
    l=[]
    for i in lst:
        l.append(i)
    return l

def split_BKT_BKC(products_and_rows,sign): #不能分离BKJ
    BK_correspondence = {}
    products_and_rows_splited = {}
    for product_name in products_and_rows:  # 此时字典如{'KLD-BKS7-400V': [11, 14, 17, 22], 'KLD-MRT16': [12, 15, 18, 23], 'KLD-AMS-4L-400V': [20, 21]}
        if "BKT" in product_name and not sign:
            BKS_in_BKT_name = product_name.replace("BKT", "BKS")
            if BKS_in_BKT_name in products_and_rows_splited:
                products_and_rows_splited[BKS_in_BKT_name] += copy_lst_from_dic(products_and_rows[product_name])
            else:
                products_and_rows_splited[BKS_in_BKT_name] = copy_lst_from_dic(products_and_rows[product_name])
            # if "不" in product_name and "电抗" in product_name: #新版本移除此功能
            #     product_name=product_name[0:product_name.index('V')+1]
            # 在products_and_rows_noBKT中加可控硅
            if "-1-" in product_name:  # KLD-BKT7-1-230V
                product_name_lst = product_name.split("-")
                if product_name_lst[3] == "230V":
                    products_and_rows_splited["KLD-TSC-1-2010X"] = []
                    BK_correspondence[product_name] = "KLD-TSC-1-2010X"
            else:
                product_name_lst = product_name.split("-")  # KLD-BKT7-400V
                if product_name_lst[2] == "400V":
                    products_and_rows_splited["KLD-TSC-3-2010X"] = []
                    BK_correspondence[product_name] = "KLD-TSC-3-2010X"
                else:
                    products_and_rows_splited["KLD-TSC-3-2010X-" + product_name_lst[2]] = []
                    BK_correspondence[product_name] = "KLD-TSC-3-2010X-" + product_name_lst[2]
        elif "BKC" in product_name and not sign:
            BKS_in_BKC_name = product_name.replace("BKC", "BKS")
            if BKS_in_BKC_name in products_and_rows_splited:
                products_and_rows_splited[BKS_in_BKC_name] += copy_lst_from_dic(products_and_rows[product_name])
            else:
                products_and_rows_splited[BKS_in_BKC_name] = copy_lst_from_dic(products_and_rows[product_name])
            if "-1-" in product_name:  # KLD-BKC7-1-230V
                product_name_lst = product_name.split("-")
                if product_name_lst[3] == "230V":
                    products_and_rows_splited["KLD-CS-1-2010"] = []
                    BK_correspondence[product_name] = "KLD-CS-1-2010"
            else:
                product_name_lst = product_name.split("-")  # KLD-BKT7-400V
                if product_name_lst[2] == "400V":
                    products_and_rows_splited["KLD-CS-3-2010"] = []
                    BK_correspondence[product_name] = "KLD-CS-3-2010"
                else:
                    products_and_rows_splited["KLD-CS-3-2010-" + product_name_lst[2]] = []
                    BK_correspondence[product_name] = "KLD-CS-3-2010-" + product_name_lst[2]
        elif "-BKS" in product_name:
            if product_name in products_and_rows_splited:
                products_and_rows_splited[product_name] += copy_lst_from_dic(products_and_rows[product_name])
            else:
                products_and_rows_splited[product_name] = products_and_rows[product_name]
        else:
            products_and_rows_splited[product_name] = products_and_rows[product_name]
    return BK_correspondence,products_and_rows_splited

def fenlu(remark: str, phase: int,sign=False): #"50kvar*6+25kvar*2",3
    remark = remark.strip()
    if remark.startswith("分路"):
        remark=remark.lstrip("分路")[1:]
    remark_lst = remark.lower().split(" ")
    for str in remark_lst:
        if "kvar" in str:
            FL_lst = str.split("+")
            for grp_num in range(len(FL_lst)):
                if "*" in FL_lst[grp_num]:
                    FL_lst[grp_num] = FL_lst[grp_num].split("*")
                else:
                    FL_lst[grp_num] = [FL_lst[grp_num], "1"]
            for num in range(len(FL_lst)):  # 将kvar调到前(并去掉“kvar”)，数量调到后
                temp = [None, None]
                for grp in FL_lst[num]:
                    if "kvar" in grp:
                        if "." in grp:
                            temp[0] = float(grp.rstrip("kvar"))
                        else:
                            temp[0] = int(grp.rstrip("kvar"))
                    else:
                        temp[1] = int(grp)
                FL_lst[num] = temp
            lst_0 = []
            lst_1 = []
            lst_2 = []
            lst_3 = []
            if phase == 3:
                for num in range(len(FL_lst)):
                    if sign == False:
                        lst_0.append(FL_lst[num])
                    else:
                        if FL_lst[num][0] < 25:
                            lst_0.append(FL_lst[num])
                        elif FL_lst[num][0] >= 25 and FL_lst[num][0]<=30:
                            lst_1.append(FL_lst[num])
                        elif FL_lst[num][0] > 30 and FL_lst[num][0] < 50:
                            lst_2.append(FL_lst[num])
                        elif FL_lst[num][0] >= 50:
                            lst_3.append(FL_lst[num])
                if sign ==False:
                    return (lst_0,) #因为else中返回的是元组
                else:
                    return lst_0,lst_1,lst_2,lst_3
            elif phase == 1:
                for num in range(len(FL_lst)):
                    if FL_lst[num][0] <= 5:
                        lst_0.append(FL_lst[num])
                    elif FL_lst[num][0] >= 10:
                        lst_1.append(FL_lst[num])
                return lst_0, lst_1

def if_in_section(section, value):
    if section == "None":  # 如excel单元格没内容，返回的是字符串“None”。。真的很扯淡
        return False
    section = str(section)
    value = int(value)
    if "≠" in section:
        if int(section.lstrip("≠")) != value:
            return 1
    elif "," in section:
        if "(" in section:
            if not int(section.split(",")[0].lstrip("(")) < value:
                return False
        elif "[" in section:
            if not int(section.split(",")[0].lstrip("[")) <= value:
                return False
        if ")" in section:
            if not int(section.split(",")[1].rstrip(")")) > value:
                return False
        elif "]" in section:
            if not int(section.split(",")[1].rstrip("]")) >= value:
                return False
        return 2
    elif "(" in section:
        if int(section.lstrip("(")) < value:
            return 3
    elif "[" in section:
        if int(section.lstrip("[")) <= value:
            return 4
    elif ")" in section:
        if int(section.rstrip(")")) > value:
            return 5
    elif "]" in section:
        if int(section.rstrip("]")) >= value:
            return 6
    elif int(section) == value:
        return 7
    else:
        return False

def get_price_value(product_name, *args):
    wb_datas = load_workbook("KLD_DATABASE.xlsx")
    if "KLD-BKS" in product_name:
        ws = wb_datas["BKS"]
        product_lst = product_name.split("-")
        if "60Hz" not in product_name:
            if "-1-" not in product_name:
                row = 1
                while ws.cell(row, 2).value != 3:
                    row += 1
                rate = int(product_lst[1].lstrip("BKS"))  # rate代表电抗率
                while ws.cell(row, 3).value != rate:
                    row += 1
                voltage = product_lst[2].rstrip("V")  # voltage代表系统电压
                while not if_in_section(str(ws.cell(row, 4).value), voltage):
                    row += 1
                return ws.cell(row + args[0], 6).value, ws.cell(row + args[0], 7).value
            elif "-1-" in product_name:
                row = 1
                while ws.cell(row, 2).value != 1:  # 不是“1”，因为若是int，openpyxl取到的是int
                    row += 1
                rate = int(product_lst[1].lstrip("BKS"))  # rate代表电抗率
                while ws.cell(row, 3).value != rate:
                    row += 1
                voltage = product_lst[3].rstrip("V")  # voltage代表系统电压
                while not if_in_section(str(ws.cell(row, 4).value), voltage):
                    row += 1
                if args[0] == 0:
                        return ws.cell(row, 6).value, ws.cell(row, 7).value
                elif args[0] == 1:
                        return ws.cell(row + 1, 6).value, ws.cell(row + 1, 7).value

    elif "KLD-TSC" in product_name:  # KLD-TSC-3-2010A-690V
        ws = wb_datas["TSC"]
        product_lst = product_name.split("-")
        if product_lst[2] == "3":
            row = 1
            while ws.cell(row, 1).value != 3:
                row += 1
            if len(product_lst) == 4:
                while ws.cell(row, 2).value != 400:
                    row += 1
                return ws.cell(row, 3).value, ws.cell(row, 4).value
            else:
                voltage=product_lst[4].rstrip("V")
                while not if_in_section(str(ws.cell(row, 2).value), voltage):#特殊电压可控硅待确认
                    row += 1
                return ws.cell(row, 3).value, ws.cell(row, 4).value
        elif product_lst[2] == "1":
            row = 1
            while ws.cell(row, 1).value != 1:
                row += 1
            if len(product_lst) == 4:
                while ws.cell(row, 2).value != 230:
                    row += 1
                return ws.cell(row, 3).value,ws.cell(row, 4).value
            else:
                voltage=product_lst[4].rstrip("V")
                while not if_in_section(str(ws.cell(row, 2).value), voltage):#特殊电压可控硅待确认
                    row += 1
                return ws.cell(row, 3).value, ws.cell(row, 4).value

    elif "KLD-AMS" in product_name:  # KLD-AMS-4L-400V-RD-G
        ws = wb_datas["AMS"]
        product_lst = product_name.split("-")
        row = 1
        if product_lst[-1] == "G":
            while ws.cell(row, 1).value != "G":
                row += 1
        while not if_in_section(str(ws.cell(row, 2).value), product_lst[3].rstrip("V")):
            row += 1
        while not if_in_section(str(ws.cell(row, 3).value), args[0]):
            row += 1
        if ws.cell(row, 4).value == None:
            return ws.cell(row, 5).value, ws.cell(row, 6).value, ws.cell(row, 7).value,0
        else:
            return ws.cell(row, 4).value, ws.cell(row, 6).value, ws.cell(row, 7).value,1

    elif "KLD-SVG" in product_name:  # KLD-SVG-4L-400V-LRD
        ws = wb_datas["SVG"]
        product_lst = product_name.split("-")
        row = 1
        if product_lst[-1] == "G":
            while ws.cell(row, 1).value != "G":
                row += 1
        if product_lst[-1].startswith("L"):
            while ws.cell(row, 1).value != "L":
                row += 1
        while not if_in_section(str(ws.cell(row, 2).value), product_lst[3].rstrip("V")):
            row += 1
        while not if_in_section(str(ws.cell(row, 3).value), args[0]):
            row += 1
        if ws.cell(row, 4).value == None:
            return ws.cell(row, 5).value, ws.cell(row, 6).value, ws.cell(row, 7).value,0
        else:
            return ws.cell(row, 4).value, ws.cell(row, 6).value, ws.cell(row, 7).value,1

    elif "MK-H1" in product_name:#KLD-MK-H1-7.2√3KV
        ws = wb_datas["MK-H1"]
        product_lst = product_name.split("-")
        if product_lst[3] == "11/√3KV" or product_lst[3] == "12/√3KV":
            row = 1
            while ws.cell(row, 1).value != 10:
                row+=1
            if args[0] == 0:
                return ws.cell(row, 3).value, ws.cell(row, 4).value
            elif args[0] == 1:
                row+=1
                return ws.cell(row, 3).value, ws.cell(row, 4).value
            elif args[0] == 2:
                row+=2
                return ws.cell(row, 3).value, ws.cell(row, 4).value
        elif product_lst[3] == "6.6/√3KV" or product_lst[3] == "7.2/√3KV":
            row = 1
            while ws.cell(row, 1).value != 6:
                row+=1
            if args[0] == 0:
                return ws.cell(row, 3).value, ws.cell(row, 4).value
            elif args[0] == 1:
                row+=1
                return ws.cell(row, 3).value, ws.cell(row, 4).value
            elif args[0] == 2:
                row+=2
                return ws.cell(row, 3).value, ws.cell(row, 4).value

    elif "FD" in product_name and "-H" in product_name:#KLD-FD6-H3-6KV-L
        ws = wb_datas["FD-H"]
        product_lst = product_name.split("-")
        voltage=float(product_lst[3].rstrip("KV"))
        phase=int(product_lst[2].lstrip("H"))
        rate=float(product_lst[1].lstrip("FD"))
        TorL=product_lst[4]
        row = 1
        while ws.cell(row, 1).value != voltage:
            row+=1
        while ws.cell(row, 2).value != phase:
            row+=1
        while ws.cell(row, 3).value != rate:
            row+=1
        while ws.cell(row, 4).value !=TorL:
            row+=1
        if args[0]==2:
            return ws.cell(row, 6).value, ws.cell(row, 7).value
        elif args[0] == 1:
            row+=1
            return ws.cell(row, 6).value, ws.cell(row, 7).value
        elif args[0] == 0:
            row+=2
            return ws.cell(row, 6).value, ws.cell(row, 7).value

    elif "KLD-CS" in product_name:
        ws = wb_datas["CS"]
        product_lst = product_name.split("-")
        row = 1
        if product_lst[2]=='3':
            while ws.cell(row, 1).value != 3:
                row+=1
            return ws.cell(row, 2).value, ws.cell(row, 3).value
        elif product_lst[2]=='1':
            while ws.cell(row, 1).value != 1:
                row+=1
            return ws.cell(row, 2).value, ws.cell(row, 3).value

    else:
        ws = wb_datas["OTHERS"]
        end_row = coordinate_from_string(ws.dimensions.split(":")[1])[1]
        for row in range(1, end_row + 1):
            if product_name == ws.cell(row, 1).value:
                return ws.cell(row, 2).value, ws.cell(row, 3).value
        return None,None