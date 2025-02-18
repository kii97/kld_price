from openpyxl.styles import Border, Side, Alignment,Font
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string, get_column_letter
from openpyxl import load_workbook

def get_the_range(ws): #返回xlsx文件的范围：初始点坐标元组，结束点坐标元组
    start_coord_str,end_coord_str = ws.dimensions.split(":")
    temp_tuple=coordinate_from_string(start_coord_str)
    start_coord=(temp_tuple[1],column_index_from_string(temp_tuple[0]))
    temp_tuple=coordinate_from_string(end_coord_str)
    end_coord=(temp_tuple[1],column_index_from_string(temp_tuple[0]))
    return start_coord,end_coord#(行，列)，（行， 列）

def find_the_projectname(ws,start_coord,end_coord):#返回第一个开头为“项目名称”的单元格的坐标
    for col in range(start_coord[1],end_coord[1]+1):
        for row in range(start_coord[0],end_coord[0]+1):
            if ws.cell(row,col).value != None and ws.cell(row,col).value.startswith("项目名称"):
                return row,col

def find_the_text(ws,start_coord,end_coord,text):
    lst=[]
    for col in range(start_coord[1],end_coord[1]+1):
        for row in range(start_coord[0],end_coord[0]+1):
            if type(ws.cell(row,col).value) == str and ws.cell(row,col).value.startswith(text):
                lst.append((row,col))
    return lst

def filtration(ws,lst):#判断“型号”右侧是否“单位”再右侧是否“数量”不符合删除
    del_lst=[]
    for i in lst:
        if ws.cell(i[0],i[1]+1).value!="单位" or ws.cell(i[0],i[1]+2).value!="数量":
            del_lst.append(i)
    for d in del_lst:
        lst.remove(d)
    return lst

def find_cell_not_merged(ws,row,col):#向左寻找合并单元格的主格，返回列
    while type(ws.cell(row, col)).__name__ == 'MergedCell':
        col-=1
    return col

def FAE_AP_get_datas(ws,lst):
    datas_dic={}
    repeat_int=0
    for c in lst:
        data_in_every_room={}
        row=c[0]
        col=c[1]
        room_name=ws.cell(row-1,find_cell_not_merged(ws,row-1,col)).value
        row+=1
        while type(ws.cell(row,col)).__name__!='MergedCell':
            if ws.cell(row,col+2).value!=0:
                if ws.cell(row,col).value == None:
                    data_in_every_room[ws.cell(row, col - 1).value] = [ws.cell(row, col - 1).value,ws.cell(row, col + 2).value]
                else:
                    data_in_every_room[ws.cell(row,col).value]=[ws.cell(row,col-1).value,ws.cell(row,col+2).value]
            row+=1
        if room_name in datas_dic:
            room_name+=str(repeat_int)
            repeat_int+=1
        datas_dic[room_name]=data_in_every_room
    return datas_dic

def FAE_AP_price_adjustment_list(ws,datas_dic,coord_str,pro_str):
    #建立调价表
    coord = [coordinate_from_string(coord_str)[1], column_index_from_string(coordinate_from_string(coord_str)[0])]
    ws.cell(coord[0], coord[1]).value = "调价表"
    ws.cell(coord[0] + 1, coord[1]).value = "涉及产品"
    ws.cell(coord[0] + 1, coord[1] + 1).value = "底价"
    ws.cell(coord[0] + 1, coord[1] + 2).value = "市场价"
    ws.cell(coord[0] + 1, coord[1] + 3).value = "底价系数"
    ws.cell(coord[0] + 1, coord[1] + 4).value = "最终底价"
    ws.cell(coord[0] + 1, coord[1] + 5).value = "调价"
    row_pdt=coord[0]+2
    price_row_dic={}
    for reg in datas_dic:#完成price_row_di的写入与data_dic内产品在调价表内行号的追加
        for pdt in datas_dic[reg]:
            if pdt in price_row_dic:
                datas_dic[reg][pdt].append(price_row_dic[pdt])#在datas_dic中追加产品在调价表的行号
            else:
                price_row_dic[pdt]=row_pdt
                datas_dic[reg][pdt].append(row_pdt)#在datas_dic中追加产品在调价表的行号
                end_row=row_pdt
                row_pdt+=1
    if pro_str=="AP":
        price_row_dic["辅材"]=row_pdt
        end_row=row_pdt
    #开始写入调价表。依据price_row_dic
    for pdt in price_row_dic:
        temp_row=price_row_dic[pdt]
        ws.cell(temp_row,coord[1]).value=pdt
        ws.cell(temp_row, coord[1]+3).value=1
        ws.cell(temp_row, coord[1]+4).value=f"={get_column_letter(coord[1]+3)}{temp_row}*{get_column_letter(coord[1]+1)}{temp_row}"
    #调整格式
    ws.merge_cells(f"{get_column_letter(coord[1])}{coord[0]}:{get_column_letter(coord[1] + 5)}{coord[0]}")
    ws.cell(coord[0],coord[1]).alignment = Alignment(horizontal='center', vertical='center')
    borders = ws[f"{get_column_letter(coord[1])}{coord[0]}:{get_column_letter(coord[1] + 5)}{end_row}"]
    thin_boader = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),bottom=Side(style='thin'))
    for j in borders:
        for i in j:
            i.border = thin_boader
    ws.column_dimensions[get_column_letter(coord[1])].width = 26
    return price_row_dic,datas_dic,coord,row_pdt

def FAE_AP_lookup_fire_price_value(pdt,pro_str):
    wb=load_workbook("KLD_DATABASE.xlsx")
    ws=wb[pro_str]
    end_row = coordinate_from_string(ws.dimensions.split(":")[1])[1]
    for row in range(1,end_row+1):
        if ws.cell(row,1).value == pdt:
            return ws.cell(row,2).value,ws.cell(row,3).value
    return None,None

def FAE_AP_write_price(ws,price_row_dic,coord,pro_str):
    for pdt in price_row_dic:
        base_price,market_price=FAE_AP_lookup_fire_price_value(pdt,pro_str)
        ws.cell(price_row_dic[pdt],coord[1]+1).value=base_price
        ws.cell(price_row_dic[pdt],coord[1]+2).value=market_price

def combine(datas_dic):
    temp_dic={}
    for reg in datas_dic:
        for pro in datas_dic[reg]:
            if pro in temp_dic:
                temp_dic[pro][1]+=datas_dic[reg][pro][1]
            else:
                temp_dic[pro]=datas_dic[reg][pro]
    return {"":temp_dic}


def fill_in_the_form(ws,datas_dic,coord,pro_str,price_row_fucai,combine_sign, start_coord_p, end_coord_p):#把产品数据填入报价单 #此处ws为报价单的ws
    row=10
    lst_AP_AS=[]
    row_MCU=0
    for reg_name in datas_dic:
        if not combine_sign:
            ws.cell(row,1).value=reg_name
            ws.cell(row,1).alignment = Alignment(horizontal='left', vertical='center')
            ws.merge_cells(f"A{row}:F{row}")
            row+=1
        if pro_str == "AP":
            ws.cell(row, 1).value ="主材"
            ws.cell(row,1).font = Font(name="宋体",size=9,bold=True)#字体
            row+=1
            num=0
        seq=1#序号
        reg_pro_rows=[]
        for pdt in datas_dic[reg_name]:
            reg_pro_rows.append(row)
            ws.cell(row,1).value=seq#填入序号
            ws.cell(row,2).value=datas_dic[reg_name][pdt][0]#填入产品名称
            ws.cell(row,3).value=pdt#填入型号
            ws.cell(row,4).value=datas_dic[reg_name][pdt][1]#填入数量
            ws.cell(row,5).value="台"#填入“台”
            ws.cell(row,6).value=f"=IF({get_column_letter(coord[1]+5)}{datas_dic[reg_name][pdt][2]},{get_column_letter(coord[1]+5)}{datas_dic[reg_name][pdt][2]},{get_column_letter(coord[1]+4)}{datas_dic[reg_name][pdt][2]})"
            ws.cell(row,7).value=f"=D{row}*F{row}"#设定公式：金额=数量*单价
            if pdt.startswith("电源模块"):
                ws.merge_cells(f"B{row}:C{row}")
            if pdt.startswith("KLD-AP-MCU"):#获取该区域主控系统的行号
                row_MCU=row
            if pdt.startswith("KLD-AP-AS"):
                lst_AP_AS.append(row)
            row+=1
            seq+=1
        if pro_str == "AP":#加入辅材
            reg_pro_rows.append(row+1)
            AP_auxiliary_material(ws, [row,1], seq, row_MCU,price_row_fucai,coord)
            row+=6
        ws.cell(row,1).value=f'="合计："&TEXT(G{row},"[dbnum2]")&"元整"'
        ws.cell(row,7).value=f'=SUM({"".join(["G"+str(i)+"," for i in reg_pro_rows]).rstrip(",")})'
        ws.merge_cells(f"A{row}:F{row}")
        ws.cell(row, 1).font = Font(name="宋体", size=10, bold=True)
        ws.cell(row, 7).font = Font(name="Times New Roman", size=10, bold=True)
        ws.cell(row, 1).alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
        row+=1

    if pro_str == "AP":#设定安装费公式
        finded_coord_lst = find_the_text(ws, start_coord_p, end_coord_p, '="安装费合计：')
        install_coord_row=finded_coord_lst[0][0]
        sum_str=""
        for r in lst_AP_AS:
            sum_str+=f",D{r}"
        sum_str=sum_str.lstrip(",")
        ws.cell(install_coord_row,7).value=f"=15000+300*SUM({sum_str})"
    return row

def date_and_name(ws,date_today,name,project_name,start_coord, end_coord):
    dt_coord=find_the_text(ws, start_coord, end_coord, "报价日期")
    nm_coord=find_the_text(ws, start_coord, end_coord, "NAME")
    pnm_coord=find_the_text(ws, start_coord, end_coord, "项目名称：")
    ws.cell(dt_coord[0][0],dt_coord[0][1]).value=f"报价日期：{date_today}"
    ws.cell(nm_coord[0][0],nm_coord[0][1]).value=f"NAME：{name}"
    ws.cell(pnm_coord[0][0],pnm_coord[0][1]).value=f"项目名称：{project_name}"

def AP_auxiliary_material(ws,coord,seq,row_MCU,price_row_fucai,coord_adj):
    row,col=coord
    ws.cell(row, col).value ="辅材"
    ws.cell(row,col).font = Font(name="宋体",size=9,bold=True)#字体
    row+=1
    ws.cell(row,col).value=seq
    ws.merge_cells(f"{get_column_letter(col)}{row}:{get_column_letter(col)}{row+4}")
    col+=1
    for t in enumerate(["数据线接头","数据传输线","光 纤","研磨包","尼龙扎带"]):
        ws.cell(row+t[0],col).value=t[1]
    col += 1
    for t in enumerate(["Data cable connector, RJ-45SH","Data transmission cable, RJ-45,","Photo electrical cable, HFBR-EUS","Grinding kit ","Nylon cable tie ,150mm"]):
        ws.cell(row+t[0],col).value=t[1]
    col += 1
    ws.cell(row, col).value = f'=D{row_MCU}&"套\n（跟主控单元数量有关）"'
    ws.merge_cells(f"{get_column_letter(col)}{row}:{get_column_letter(col)}{row + 4}")
    ws.cell(row,col).font = Font(name = "宋体", size = 9)
    ws.cell(row,col).alignment = Alignment(wrap_text=True,horizontal='center', vertical='center')
    col+=1
    for t in enumerate(["个","m","m","个","个"]):
        ws.cell(row+t[0],col).value=t[1]
        if t[0]==1 or t[0] ==2:
            ws.cell(row + t[0], col).font=Font(name="Times New Roman",size=9)
    col+=1
    ws.cell(row,col).value=f"=IF({get_column_letter(coord_adj[1]+5)}{price_row_fucai},{get_column_letter(coord_adj[1]+5)}{price_row_fucai},{get_column_letter(coord_adj[1]+4)}{price_row_fucai})"
    ws.merge_cells(f"{get_column_letter(col)}{row}:{get_column_letter(col)}{row + 4}")
    col+=1
    ws.merge_cells(f"{get_column_letter(col)}{row}:{get_column_letter(col)}{row + 4}")
    ws.cell(row, 7).value = f"=D{row_MCU}*F{row}"  # 设定公式：金额=数量*单价